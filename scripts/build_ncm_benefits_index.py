#!/usr/bin/env python3
"""Build an NCM x tax-benefit index from validated legal excerpts.

The output is intentionally strict: a row only exists when an NCM/TIPI-style
code appears in a legal excerpt that also contains a tax benefit/treatment.
"""

from __future__ import annotations

import csv
import hashlib
import json
import os
import re
import sys
import time
import urllib.request
from collections import Counter, defaultdict
from datetime import date
from functools import lru_cache
from pathlib import Path


ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / "scripts"
sys.path.insert(0, str(SCRIPTS))

from state_legal_pages import STATE_NAMES, collect_state_documents  # noqa: E402
from validated_benefits import (  # noqa: E402
    FEDERAL_FILES,
    PlainTextHTMLParser,
    build_entry,
    compact,
    federal_source_meta,
    goias_source_meta,
    legal_windows,
    load_json,
    official_state_sources,
    source_meta_from_doc,
    table_line_windows,
)


OUT_JSON = ROOT / "data" / "ncm_benefits_index.json"
OUT_CSV = ROOT / "data" / "ncm_benefits_index.csv"
CONFAZ_INDEX = ROOT / "data" / "confaz_ultimos_5_anos.json"
TIPI_XLSX = Path(r"C:\Users\kris2\OneDrive\COWORK\BD_LEGISLACAO\#FEDERAIS-COMPILADO-ONLINE\TIPI_Vigente_2022.xlsx")
TODAY = date.today().isoformat()

STATE_SOURCE_CATEGORIES = {
    "RICMS",
    "ICMS_LEIS",
    "ICMS_DECRETOS",
    "ICMS_BENEFICIOS",
    "ICMS_ANEXOS",
    "ICMS_ST",
    "ICMS_ALIQUOTAS",
    "INSTRUCOES_NORMATIVAS",
    "PORTARIAS",
    "RESOLUCOES",
    "OUTROS",
}

NCM_CONTEXT_RE = re.compile(
    r"(NCM|TIPI|NBM|Nomenclatura|posi[cç][aã]o|posi[cç][oõ]es|classificad[oa]s?|c[oó]digo(?:s)?\s+(?:da\s+)?NCM)",
    re.I,
)
NCM_CODE_RE = re.compile(
    r"(?<!\d)(\d{4}\.\d{2}\.\d{2}|\d{4}\.\d{2}|\d{2}\.\d{2}|\d{8}|\d{4})(?!\d)"
)


def clean_text(value: str) -> str:
    return " ".join((value or "").split())


def ncm_digits(code: str) -> str:
    return re.sub(r"\D", "", code or "")


def ncm_level(code: str) -> str:
    digits = ncm_digits(code)
    if len(digits) == 8:
        return "subitem NCM"
    if len(digits) == 6:
        return "subposição NCM"
    if len(digits) == 4:
        return "posição NCM"
    if len(digits) == 2:
        return "capítulo NCM"
    return "NCM"


@lru_cache(maxsize=1)
def valid_ncm_prefixes() -> set[str]:
    prefixes: set[str] = set()
    if not TIPI_XLSX.exists():
        return prefixes
    try:
        from openpyxl import load_workbook
    except Exception:
        return prefixes
    workbook = load_workbook(TIPI_XLSX, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    for row in sheet.iter_rows(values_only=True):
        value = str(row[0] or "").strip()
        if not value:
            continue
        digits = ncm_digits(value)
        if len(digits) < 4 or set(digits) == {"0"}:
            continue
        for size in (4, 6, 8):
            if len(digits) >= size:
                prefixes.add(digits[:size])
    workbook.close()
    return prefixes


def is_valid_ncm(digits: str) -> bool:
    prefixes = valid_ncm_prefixes()
    if not prefixes:
        return True
    return digits in prefixes


def strict_ncms(excerpt: str) -> list[str]:
    """Extract NCM-like codes only when the surrounding text says NCM/TIPI/etc."""
    text = clean_text(excerpt)
    found: list[str] = []
    seen: set[str] = set()
    for context in NCM_CONTEXT_RE.finditer(text):
        start = max(0, context.start() - 90)
        end = min(len(text), context.end() + 260)
        window = text[start:end]
        for match in NCM_CODE_RE.finditer(window):
            code = match.group(1)
            digits = ncm_digits(code)
            if len(digits) not in {4, 6, 8}:
                continue
            if set(digits) == {"0"} or digits.startswith("00"):
                continue
            if not is_valid_ncm(digits):
                continue
            if len(digits) == 4 and 1900 <= int(digits) <= 2099 and "." not in code:
                continue
            key = digits
            if key not in seen:
                seen.add(key)
                found.append(code)
    return found[:30]


def usable_state_doc(doc: dict) -> bool:
    category = doc.get("category", "")
    if category not in STATE_SOURCE_CATEGORIES:
        return False
    flags = " ".join(doc.get("flags", [])).lower()
    if "escopo dominante incompat" in flags or "fallback amplo contaminado" in flags:
        return False
    return True


def all_state_sources() -> list[dict]:
    manifests = official_state_sources()
    sources: list[dict] = []
    goias = goias_source_meta()
    if goias:
        sources.append(goias)
    for uf in sorted(STATE_NAMES):
        if uf == "GO":
            continue
        for doc in collect_state_documents(uf):
            if not usable_state_doc(doc):
                continue
            try:
                meta = source_meta_from_doc(uf, doc, manifests.get(uf, {}))
            except Exception:
                continue
            if (
                meta.get("official_url", "").startswith(("http://", "https://"))
                and meta.get("text")
                and NCM_CONTEXT_RE.search(meta["text"])
            ):
                sources.append(meta)
    return sources


def federal_sources() -> list[dict]:
    sources: list[dict] = []
    for config in FEDERAL_FILES:
        meta = federal_source_meta(config)
        if meta and NCM_CONTEXT_RE.search(meta.get("text", "")):
            sources.append(meta)
    return sources


def fetch_official_text(url: str) -> str:
    request = urllib.request.Request(url, headers={"User-Agent": "RJC-Conhecimento/2.0"})
    with urllib.request.urlopen(request, timeout=35) as response:
        raw = response.read()
    html = raw.decode("utf-8", errors="ignore")
    if html.count("�") > 20:
        html = raw.decode("latin-1", errors="ignore")
    parser = PlainTextHTMLParser()
    parser.feed(html)
    return parser.text()


def confaz_sources() -> list[dict]:
    payload = load_json(CONFAZ_INDEX, {})
    sources: list[dict] = []
    for family in payload.get("families", {}).values():
        for year in family.get("years", []):
            for act in year.get("acts", []):
                url = act.get("url", "")
                if not url:
                    continue
                try:
                    text = fetch_official_text(url)
                except Exception:
                    continue
                if not text or not NCM_CONTEXT_RE.search(text):
                    continue
                title = act.get("title") or url.rsplit("/", 1)[-1]
                sources.append({
                    "jurisdiction": "CONFAZ",
                    "name": "CONFAZ",
                    "tax": "ICMS",
                    "title": title,
                    "official_url": url,
                    "captured_on": TODAY,
                    "source_file": title,
                    "source_path": url,
                    "sha256": hashlib.sha256(text.encode("utf-8")).hexdigest(),
                    "text": text,
                })
                time.sleep(0.05)
    return sources


def entry_rows(source: dict) -> list[dict]:
    rows: list[dict] = []
    excerpts = legal_windows(source["text"]) + table_line_windows(source["text"])
    for seq, excerpt in enumerate(excerpts, start=1):
        ncm_codes = strict_ncms(excerpt)
        if not ncm_codes:
            continue
        entry = build_entry(source, excerpt, seq)
        if not entry:
            continue
        for code in ncm_codes:
            digits = ncm_digits(code)
            if not digits:
                continue
            row_id = hashlib.sha1(
                "|".join([
                    source["jurisdiction"],
                    source["source_file"],
                    digits,
                    entry["benefit_type"],
                    entry["legal_basis"],
                    entry["legal_excerpt"][:180],
                ]).encode("utf-8")
            ).hexdigest()[:16]
            rows.append({
                "id": f"ncm-{row_id}",
                "ncm": code,
                "ncm_digits": digits,
                "ncm_level": ncm_level(code),
                "jurisdiction": source["jurisdiction"],
                "name": source["name"],
                "tax": source["tax"],
                "origin": "CONFAZ" if source["jurisdiction"] == "CONFAZ" else ("Federal" if source["jurisdiction"] == "Federal" else "Estado"),
                "benefit_group": entry["benefit_group"],
                "benefit_type": entry["benefit_type"],
                "product_or_operation": entry["product_or_operation"],
                "conditions": entry["conditions"],
                "prohibitions": entry["prohibitions"],
                "validity_start": entry.get("validity_start", ""),
                "validity_end": entry.get("validity_end", ""),
                "transition_status": entry.get("transition_status", ""),
                "legal_nature": entry.get("legal_nature", ""),
                "proof_required": entry["proof_required"],
                "risk": entry["risk"],
                "legal_basis": entry["legal_basis"],
                "legal_excerpt": compact(entry["legal_excerpt"], 650),
                "source_title": source["title"],
                "official_url": source["official_url"],
                "captured_on": source.get("captured_on", ""),
                "sha256": source["sha256"],
            })
    return rows


def build_index(include_confaz: bool = True) -> dict:
    sources = all_state_sources() + federal_sources()
    if include_confaz:
        sources += confaz_sources()
    rows: list[dict] = []
    seen: set[str] = set()
    for source in sources:
        for row in entry_rows(source):
            key = "|".join([
                row["ncm_digits"],
                row["jurisdiction"],
                row["benefit_type"],
                clean_text(row["product_or_operation"])[:180],
                row["legal_basis"],
            ])
            if key in seen:
                continue
            seen.add(key)
            rows.append(row)
    rows.sort(key=lambda item: (item["ncm_digits"], item["jurisdiction"], item["benefit_type"], item["legal_basis"]))
    by_origin = Counter(row["origin"] for row in rows)
    by_jurisdiction = Counter(row["jurisdiction"] for row in rows)
    by_group = Counter(row["benefit_group"] for row in rows)
    return {
        "schema": "rjc-ncm-benefits-index-v1",
        "generated_on": TODAY,
        "source_rule": "cada linha contem NCM extraido de trecho legal com tratamento tributario e fonte oficial",
        "summary": {
            "rows": len(rows),
            "unique_ncm": len({row["ncm_digits"] for row in rows}),
            "jurisdictions": len({row["jurisdiction"] for row in rows}),
            "origins": dict(sorted(by_origin.items())),
            "by_jurisdiction": dict(sorted(by_jurisdiction.items())),
            "top_groups": dict(by_group.most_common(15)),
        },
        "rows": rows,
    }


def merge_cached_confaz_rows(payload: dict) -> dict:
    previous = load_json(OUT_JSON, {"rows": []})
    cached = [
        row for row in previous.get("rows", [])
        if row.get("origin") == "CONFAZ" and row.get("official_url", "").startswith("https://www.confaz.fazenda.gov.br/")
    ]
    seen = {
        "|".join([
            row.get("ncm_digits", ""),
            row.get("jurisdiction", ""),
            row.get("benefit_type", ""),
            clean_text(row.get("product_or_operation", ""))[:180],
            row.get("legal_basis", ""),
        ])
        for row in payload.get("rows", [])
    }
    for row in cached:
        key = "|".join([
            row.get("ncm_digits", ""),
            row.get("jurisdiction", ""),
            row.get("benefit_type", ""),
            clean_text(row.get("product_or_operation", ""))[:180],
            row.get("legal_basis", ""),
        ])
        if key not in seen:
            seen.add(key)
            payload["rows"].append(row)
    payload["rows"].sort(key=lambda item: (item["ncm_digits"], item["jurisdiction"], item["benefit_type"], item["legal_basis"]))
    by_origin = Counter(row["origin"] for row in payload["rows"])
    by_jurisdiction = Counter(row["jurisdiction"] for row in payload["rows"])
    by_group = Counter(row["benefit_group"] for row in payload["rows"])
    payload["summary"] = {
        "rows": len(payload["rows"]),
        "unique_ncm": len({row["ncm_digits"] for row in payload["rows"]}),
        "jurisdictions": len({row["jurisdiction"] for row in payload["rows"]}),
        "origins": dict(sorted(by_origin.items())),
        "by_jurisdiction": dict(sorted(by_jurisdiction.items())),
        "top_groups": dict(by_group.most_common(15)),
        "confaz_mode": "cache_reutilizado",
    }
    return payload


def write_csv(payload: dict) -> None:
    OUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    fields = [
        "ncm",
        "ncm_digits",
        "ncm_level",
        "origin",
        "jurisdiction",
        "tax",
        "benefit_group",
        "benefit_type",
        "product_or_operation",
        "conditions",
        "prohibitions",
        "validity_start",
        "validity_end",
        "transition_status",
        "legal_nature",
        "proof_required",
        "risk",
        "legal_basis",
        "source_title",
        "official_url",
        "legal_excerpt",
    ]
    with OUT_CSV.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(payload["rows"])


def main() -> int:
    refresh_confaz = os.environ.get("RJC_REFRESH_CONFAZ_NCM", "").strip() == "1"
    payload = build_index(include_confaz=refresh_confaz)
    if not refresh_confaz:
        payload = merge_cached_confaz_rows(payload)
    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    write_csv(payload)
    print(json.dumps(payload["summary"], ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
