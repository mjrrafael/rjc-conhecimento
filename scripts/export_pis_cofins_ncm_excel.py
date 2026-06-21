#!/usr/bin/env python3
"""Export the public PIS/Cofins NCM dataset to a styled searchable Excel file."""

from __future__ import annotations

import json
import os
import sys
from collections import Counter
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
NDJSON = ROOT / "data" / "pis-cofins" / "ncm.ndjson"
INDEX = ROOT / "data" / "pis-cofins" / "ncm-index.json"
DEFAULT_OUT = Path(
    os.environ.get(
        "RJC_PIS_COFINS_NCM_EXCEL",
        rf"G:\Meu Drive\RJC\BD_LEGISLACAO\PIS_COFINS_NCM\pis-cofins-ncm-{date.today().isoformat()}.xlsx",
    )
)

HEAD_FILL = PatternFill("solid", fgColor="16365F")
SUB_FILL = PatternFill("solid", fgColor="EAF4FB")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
OK_FILL = PatternFill("solid", fgColor="E8F4EF")
WHITE = Font(color="FFFFFF", bold=True)
NAVY = Font(color="16365F", bold=True)
MUTED = Font(color="5F6B7A")
THIN = Side(style="thin", color="D8DEE8")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


FIELDS: list[tuple[str, str]] = [
    ("id", "ID"),
    ("ncm.codigo", "NCM"),
    ("ncm.digitos", "NCM digitos"),
    ("ncm.nivel", "Nivel"),
    ("tratamento", "Tratamento"),
    ("setor", "Setor"),
    ("aplicacao", "Aplicacao"),
    ("operacao", "Operacao"),
    ("etapa_cadeia", "Etapa cadeia"),
    ("status", "Status"),
    ("publicacao", "Publicacao"),
    ("inicio_vigencia", "Inicio vigencia"),
    ("inicio_eficacia", "Inicio eficacia"),
    ("fim_vigencia", "Fim vigencia"),
    ("ato_oficial.tipo", "Ato tipo"),
    ("ato_oficial.numero", "Ato numero"),
    ("ato_oficial.titulo", "Ato titulo"),
    ("ato_oficial.url", "Fonte oficial"),
    ("resumo_operacional", "Resumo operacional"),
    ("mercadoria_servico", "Mercadoria/servico"),
    ("condicoes", "Condicoes"),
    ("vedacoes", "Vedacoes"),
    ("prova_documental", "Prova documental"),
    ("risco", "Risco"),
    ("transicao_cbs.status", "Transicao CBS"),
    ("transicao_cbs.referencia", "Referencia CBS"),
    ("trecho_legal", "Trecho legal"),
    ("leitura_humana.como_validar", "Como validar"),
    ("leitura_humana.nao_usar_sem", "Nao usar sem"),
    ("verificado_em", "Verificado em"),
]


def load_ndjson(path: Path) -> list[dict[str, Any]]:
    return [json.loads(line) for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]


def get_nested(row: dict[str, Any], key: str) -> Any:
    current: Any = row
    for part in key.split("."):
        if not isinstance(current, dict):
            return ""
        current = current.get(part, "")
    if isinstance(current, list):
        return "; ".join(str(item) for item in current if str(item).strip())
    if current is None:
        return ""
    return current


def clean(value: Any) -> str:
    return " ".join(str(value or "").split())


def search_blob(row: dict[str, Any]) -> str:
    parts = [get_nested(row, key) for key, _label in FIELDS]
    parts.append(row.get("pesquisa_texto", ""))
    return clean(" ".join(str(part) for part in parts))


def style_range(ws, start_row: int, end_row: int, max_col: int) -> None:
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, max_col=max_col):
        for cell in row:
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def write_headers(ws, row: int, labels: list[str]) -> None:
    for col, label in enumerate(labels, start=1):
        cell = ws.cell(row=row, column=col, value=label)
        cell.fill = HEAD_FILL
        cell.font = WHITE
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def autosize(ws, max_width: int = 52) -> None:
    for column_cells in ws.columns:
        letter = column_cells[0].column_letter
        width = min(max(len(clean(cell.value)) for cell in column_cells if cell.value is not None) + 2, max_width)
        ws.column_dimensions[letter].width = max(width, 12)


def add_table(ws, name: str, ref: str) -> None:
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def build_base_sheet(wb: Workbook, rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Base_NCM")
    headers = [label for _key, label in FIELDS] + ["busca_normalizada"]
    write_headers(ws, 1, headers)
    url_col = next(idx for idx, (_key, label) in enumerate(FIELDS, start=1) if label == "Fonte oficial")
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, (key, _label) in enumerate(FIELDS, start=1):
            value = get_nested(row, key)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx == url_col and value:
                cell.hyperlink = str(value)
                cell.style = "Hyperlink"
        ws.cell(row=row_idx, column=len(FIELDS) + 1, value=search_blob(row))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
    add_table(ws, "TabelaPISCOFINSNCM", f"A1:{get_column_letter(len(headers))}{len(rows) + 1}")
    style_range(ws, 1, len(rows) + 1, len(headers))
    status_col = next(idx for idx, (_key, label) in enumerate(FIELDS, start=1) if label == "Status")
    col_letter = get_column_letter(status_col)
    ws.conditional_formatting.add(
        f"{col_letter}2:{col_letter}{len(rows) + 1}",
        CellIsRule(operator="equal", formula=['"vigente"'], fill=OK_FILL),
    )
    ws.column_dimensions[get_column_letter(len(headers))].hidden = True
    autosize(ws)


def build_search_sheet(wb: Workbook, rows: list[dict[str, Any]]) -> None:
    ws = wb.active
    ws.title = "Pesquisar"
    visible_cols = len(FIELDS)
    last_visible = get_column_letter(visible_cols)
    search_col = get_column_letter(visible_cols + 1)
    last_row = len(rows) + 1

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    ws["A1"] = "PIS/Cofins por NCM - pesquisa local"
    ws["A1"].font = Font(color="16365F", bold=True, size=18)
    ws["A2"] = "Digite um NCM, descricao, setor, tratamento, ato ou termo legal em B3. A tabela abaixo usa FILTER do Excel 365; se sua versao nao suportar, use a aba Base_NCM com filtros."
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws["A3"] = "Pesquisa"
    ws["A3"].font = NAVY
    ws["B3"] = ""
    ws["B3"].fill = WARN_FILL
    ws["B3"].border = Border(bottom=Side(style="medium", color="A96516"))
    ws["B3"].alignment = Alignment(wrap_text=False)
    ws["D3"] = "Linhas publicadas"
    ws["D3"].font = NAVY
    ws["E3"] = len(rows)
    ws["D4"] = "Status"
    ws["D4"].font = NAVY
    ws["E4"] = "validado; conferir fonte antes de aplicar"
    write_headers(ws, 6, [label for _key, label in FIELDS])
    if rows:
        ws["A7"] = (
            f'=FILTER(Base_NCM!A2:{last_visible}{last_row},'
            f'ISNUMBER(SEARCH($B$3,Base_NCM!{search_col}2:{search_col}{last_row})),'
            '"Sem resultado para a busca")'
        )
    ws.freeze_panes = "A7"
    style_range(ws, 6, 6, visible_cols)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 26


def write_counter_block(ws, title: str, start_row: int, counter: Counter[str]) -> int:
    ws.cell(row=start_row, column=1, value=title).font = Font(color="16365F", bold=True, size=13)
    write_headers(ws, start_row + 1, ["Categoria", "Linhas"])
    for offset, (key, count) in enumerate(counter.most_common(), start=start_row + 2):
        ws.cell(row=offset, column=1, value=key)
        ws.cell(row=offset, column=2, value=count)
    end_row = start_row + 1 + len(counter)
    style_range(ws, start_row + 1, end_row, 2)
    return end_row + 3


def build_summary_sheet(wb: Workbook, rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Resumo")
    ws["A1"] = "Resumo da base PIS/Cofins por NCM"
    ws["A1"].font = Font(color="16365F", bold=True, size=16)
    ws["A3"] = "Linhas publicadas"
    ws["B3"] = len(rows)
    ws["A4"] = "NCM/codigos unicos"
    ws["B4"] = len({get_nested(row, "ncm.digitos") for row in rows})
    ws["A5"] = "Verificacao mais antiga"
    ws["B5"] = min((str(row.get("verificado_em", "")) for row in rows if row.get("verificado_em")), default="")
    for row_idx in range(3, 6):
        ws.cell(row=row_idx, column=1).font = NAVY
        ws.cell(row=row_idx, column=1).fill = SUB_FILL
        ws.cell(row=row_idx, column=1).border = BORDER
        ws.cell(row=row_idx, column=2).border = BORDER
    next_row = 8
    next_row = write_counter_block(ws, "Por tratamento", next_row, Counter(str(row.get("tratamento", "")) for row in rows))
    next_row = write_counter_block(ws, "Por setor", next_row, Counter(str(row.get("setor", "")) for row in rows))
    next_row = write_counter_block(ws, "Por fonte", next_row, Counter(str(row.get("source_id", "")) for row in rows))
    write_counter_block(ws, "Por status", next_row, Counter(str(row.get("status", "")) for row in rows))
    autosize(ws)


def build_sources_sheet(wb: Workbook, rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Fontes")
    headers = ["Fonte", "Ato", "Titulo", "URL oficial", "Linhas", "Publicacao", "Inicio eficacia", "Fim vigencia"]
    write_headers(ws, 1, headers)
    counter = Counter(str(row.get("source_id", "")) for row in rows)
    first_by_source: dict[str, dict[str, Any]] = {}
    for row in rows:
        first_by_source.setdefault(str(row.get("source_id", "")), row)
    for idx, (source_id, count) in enumerate(counter.most_common(), start=2):
        row = first_by_source[source_id]
        ato = row.get("ato_oficial", {}) if isinstance(row.get("ato_oficial"), dict) else {}
        values = [
            source_id,
            f"{ato.get('tipo', '')} {ato.get('numero', '')}".strip(),
            ato.get("titulo", ""),
            ato.get("url", ""),
            count,
            row.get("publicacao", ""),
            row.get("inicio_eficacia", ""),
            row.get("fim_vigencia") or "sem fim indicado",
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=idx, column=col, value=value)
            if col == 4 and value:
                cell.hyperlink = str(value)
                cell.style = "Hyperlink"
    add_table(ws, "TabelaFontesPISCOFINS", f"A1:H{len(counter) + 1}")
    style_range(ws, 1, len(counter) + 1, len(headers))
    autosize(ws)


def build_method_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Metodo")
    notes = [
        ["Regra mestre", "NCM aponta caminho, mas nao decide sozinho. Aplicar apenas com fonte oficial, vigencia, etapa da cadeia, documento fiscal e EFD-Contribuicoes."],
        ["Origem", f"Base publica: {NDJSON}"],
        ["Indice", f"Indice publico: {INDEX}"],
        ["Pesquisa", "A aba Pesquisar usa o texto normalizado da Base_NCM. A busca e auxiliar; a prova esta no ato oficial e no trecho legal."],
        ["Ressalva", "A base publica contem registros validados no lote atual. Itens em quarentena ou fontes ainda A VALIDAR nao aparecem como fato publicado."],
        ["Atualizacao", f"Arquivo gerado em {date.today().isoformat()} pelo script scripts/export_pis_cofins_ncm_excel.py."],
    ]
    write_headers(ws, 1, ["Tema", "Descricao"])
    for row_idx, values in enumerate(notes, start=2):
        ws.cell(row=row_idx, column=1, value=values[0])
        ws.cell(row=row_idx, column=2, value=values[1])
    style_range(ws, 1, len(notes) + 1, 2)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 110
    for row in ws.iter_rows(min_row=2, max_row=len(notes) + 1, max_col=2):
        row[0].font = NAVY
        row[1].alignment = Alignment(wrap_text=True, vertical="top")


def main(argv: list[str]) -> int:
    out = Path(argv[1]) if len(argv) > 1 else DEFAULT_OUT
    rows = [row for row in load_ndjson(NDJSON) if row.get("publishable") is True]
    if not rows:
        print("no public PIS/Cofins NCM rows to export")
        return 1
    out.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    build_search_sheet(wb, rows)
    build_base_sheet(wb, rows)
    build_summary_sheet(wb, rows)
    build_sources_sheet(wb, rows)
    build_method_sheet(wb)
    wb.save(out)
    print(json.dumps({"status": "OK", "rows": len(rows), "xlsx": str(out)}, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
