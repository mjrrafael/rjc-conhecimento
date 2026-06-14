#!/usr/bin/env python3
"""Extract Reforma Tributaria tables and technical notes into versioned text."""

from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader


ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "data" / "legal_sources" / "reforma_tributaria"
ORIG_DIR = OUT_DIR / "originais"

DOCS_PAGE = "https://dfe-portal.svrs.rs.gov.br/Nfe/Documentos"
CFF_CLASS = "https://dfe-portal.svrs.rs.gov.br/CFF/ClassificacaoTributaria"
CFF_CRED = "https://dfe-portal.svrs.rs.gov.br/CFF/TabelaCreditoPresumido"


def normalize_space(value: str) -> str:
    value = value.replace("\xa0", " ")
    value = re.sub(r"[ \t]+", " ", value)
    value = re.sub(r" *\n *", "\n", value)
    value = re.sub(r"\n{3,}", "\n\n", value)
    return value.strip()


def cell_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value).replace(".", ",")
    text = str(value)
    text = normalize_space(text)
    text = text.replace("|", "/")
    text = text.replace("\r", "\n").replace("\n", " / ")
    return text


def visible_rows(ws) -> list[list[str]]:
    raw_rows = [[cell_text(value) for value in row] for row in ws.iter_rows(values_only=True)]
    while raw_rows and not any(raw_rows[-1]):
        raw_rows.pop()
    if not raw_rows:
        return []
    width = max(len(row) for row in raw_rows)
    for row in raw_rows:
        row.extend([""] * (width - len(row)))
    keep_cols = []
    for index in range(width):
        if any(row[index] for row in raw_rows):
            keep_cols.append(index)
    return [[row[index] for index in keep_cols] for row in raw_rows]


def markdown_table(rows: list[list[str]]) -> str:
    if not rows:
        return ""
    width = len(rows[0])
    header = rows[0]
    body = rows[1:]
    lines = [
        "| " + " | ".join(header) + " |",
        "| " + " | ".join(["---"] * width) + " |",
    ]
    for row in body:
        row = row[:width] + [""] * max(0, width - len(row))
        lines.append("| " + " | ".join(row[:width]) + " |")
    return "\n".join(lines)


def workbook_to_text(path: Path, title: str, source_url: str, intro: list[str]) -> str:
    wb = load_workbook(path, read_only=True, data_only=True)
    parts = [
        f"# {title}",
        "",
        f"Fonte de publicacao: {DOCS_PAGE}",
        f"Consulta online: {source_url}",
        f"Arquivo incorporado ao portal: {path.name}",
        "",
        *intro,
        "",
    ]
    for ws in wb.worksheets:
        rows = visible_rows(ws)
        if not rows:
            continue
        parts.extend(
            [
                f"## Planilha: {ws.title}",
                f"Total de registros: {max(len(rows) - 1, 0)}",
                "",
                markdown_table(rows),
                "",
            ]
        )
    return "\n".join(parts).strip() + "\n"


def pdf_text(path: Path) -> str:
    reader = PdfReader(str(path))
    pages = []
    for index, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        text = normalize_space(text)
        if text:
            pages.append(f"## Pagina {index}\n{text}")
    return "\n\n".join(pages)


def write_pdf_source(path: Path, title: str, download_ref: str, intro: list[str], extra: str = "") -> str:
    parts = [
        f"# {title}",
        "",
        f"Fonte de publicacao: {DOCS_PAGE}",
        f"Arquivo: {path.name}",
        f"Referencia de download no portal: {download_ref}",
        "",
        *intro,
        "",
    ]
    if extra:
        parts.extend([extra.strip(), ""])
    parts.extend(["## Texto integral extraido do documento", pdf_text(path)])
    return "\n".join(parts).strip() + "\n"


def main() -> int:
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    cclass = workbook_to_text(
        ORIG_DIR / "cClassTrib_2026_04_15.xlsx",
        "Tabela CST e cClassTrib do IBS e da CBS - 15/04/2026",
        CFF_CLASS,
        [
            "A tabela liga o CST-IBS/CBS ao codigo de classificacao tributaria da operacao.",
            "Na prática, o CST indica a família de tratamento; o cClassTrib aponta a hipótese legal e operacional que deverá aparecer no documento fiscal eletrônico.",
        ],
    )
    (OUT_DIR / "Tabela_CST_cClassTrib_IBS_CBS_2026_04_15.txt").write_text(cclass, encoding="utf-8")

    ccred = workbook_to_text(
        ORIG_DIR / "cCredPres_2025_12_12Public.xlsx",
        "Tabela de códigos de crédito presumido do IBS e da CBS - 12/12/2025",
        CFF_CRED,
        [
            "A tabela estrutura as hipóteses legais de crédito presumido da LC 214/2025.",
            "Ela indica se a apropriação ocorre no documento fiscal ou por evento, qual grupo deve ser preenchido e qual regra de alíquota ou percentual se aplica.",
        ],
    )
    (OUT_DIR / "Tabela_cCredPres_IBS_CBS_2025_12_12.txt").write_text(ccred, encoding="utf-8")

    aliquotas = """
## Quadro operacional - alíquotas padrão do IBS e da CBS
| Ano | pIBSUF (%) | pIBSMun (%) | pCBS (%) | Leitura prática |
| --- | --- | --- | --- | --- |
| 2026 | 0,1 | 0 | 0,9 | Ano de teste nos documentos fiscais, com CBS e IBS informados conforme a LC 214/2025 e o Informe Técnico. |
| 2027 | 0,05 | 0,05 | Aguardar legislação | Cada ente deverá definir suas alíquotas por lei própria; na ausência, aplica-se a alíquota de referência conforme a LC 214/2025. |
| 2028 | 0,05 | 0,05 | Aguardar legislação | Mantém-se a lógica de teste/transição para IBS, aguardando disciplina específica da CBS. |
| 2029 em diante | Aguardar legislação | Aguardar legislação | Aguardar legislação | A operação passa a depender das alíquotas fixadas pelos entes ou da alíquota de referência aplicável. |
"""
    it_text = write_pdf_source(
        ORIG_DIR / "IT_2025_002_v1_50_Tabelas_Classificacao_IBS_CBS.pdf",
        "Informe Tecnico 2025.002 v1.50 - tabelas de classificacao do IBS e da CBS",
        "download_arquivo_estatico('DFE', 16, 'IT 2025.002 v.1.50 - Tabelas de Classificacao do IBS e da CBS.pdf')",
        [
            "O informe técnico explica as tabelas CST, cClassTrib, cCredPres e o quadro de alíquotas padrão a serem informadas nos documentos fiscais eletrônicos.",
        ],
        extra=aliquotas,
    )
    (OUT_DIR / "IT_2025_002_v1_50_Tabelas_Classificacao_IBS_CBS.txt").write_text(it_text, encoding="utf-8")

    nt_text = write_pdf_source(
        ORIG_DIR / "NT_2025_002_v1_35_RTC_NFe_IBS_CBS_IS.pdf",
        "Nota Tecnica 2025.002 v1.35 - referencia historica superada para NF-e/NFC-e",
        "download_arquivo_estatico('NFE', 3, 'NT_2025.002_v1.35_RTC_NF-e_IBS_CBS_IS.pdf')",
        [
            "A versão 1.35 deve permanecer marcada como referência histórica, pois foi superada por versões posteriores da NT 2025.002.",
            "Não utilizar esta versão como orientação operacional vigente sem captura e validação da versão vigente publicada no Portal NF-e.",
        ],
    )
    (OUT_DIR / "NT_2025_002_v1_35_RTC_NFe_IBS_CBS_IS.txt").write_text(nt_text, encoding="utf-8")

    print("Fontes da Reforma extraidas para data/legal_sources/reforma_tributaria")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
