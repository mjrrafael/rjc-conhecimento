#!/usr/bin/env python3
"""Audit the exported PIS/Cofins NCM Excel workbook."""

from __future__ import annotations

import json
import os
import sys
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
NDJSON = ROOT / "data" / "pis-cofins" / "ncm.ndjson"
DEFAULT_XLSX = Path(
    os.environ.get(
        "RJC_PIS_COFINS_NCM_EXCEL",
        rf"G:\Meu Drive\RJC\BD_LEGISLACAO\PIS_COFINS_NCM\pis-cofins-ncm-{date.today().isoformat()}.xlsx",
    )
)


def load_ndjson(path: Path) -> list[dict]:
    return [json.loads(line) for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]


def main(argv: list[str]) -> int:
    xlsx = Path(argv[1]) if len(argv) > 1 else DEFAULT_XLSX
    errors: list[str] = []
    rows = [row for row in load_ndjson(NDJSON) if row.get("publishable") is True]
    if not xlsx.exists():
        errors.append(f"workbook not found: {xlsx}")
    else:
        wb = load_workbook(xlsx, data_only=False)
        expected_sheets = {"Pesquisar", "Base_NCM", "Resumo", "Fontes", "Metodo"}
        missing = expected_sheets - set(wb.sheetnames)
        if missing:
            errors.append(f"missing sheets: {', '.join(sorted(missing))}")
        if "Base_NCM" in wb.sheetnames:
            base = wb["Base_NCM"]
            if base.max_row != len(rows) + 1:
                errors.append(f"Base_NCM rows {base.max_row - 1} differ from NDJSON {len(rows)}")
            headers = [str(base.cell(row=1, column=col).value or "") for col in range(1, base.max_column + 1)]
            for required in ["ID", "NCM", "Tratamento", "Resumo operacional", "Como validar", "Nao usar sem", "busca_normalizada"]:
                if required not in headers:
                    errors.append(f"Base_NCM missing header: {required}")
            search_col = headers.index("busca_normalizada") + 1 if "busca_normalizada" in headers else 0
            if search_col and not base.column_dimensions[get_column_letter(search_col)].hidden:
                errors.append("Base_NCM search helper column must be hidden")
            ids = {str(base.cell(row=row_idx, column=1).value or "") for row_idx in range(2, base.max_row + 1)}
            ndjson_ids = {str(row.get("id", "")) for row in rows}
            if ids != ndjson_ids:
                errors.append("Base_NCM IDs differ from NDJSON IDs")
        if "Pesquisar" in wb.sheetnames:
            search = wb["Pesquisar"]
            formula = str(search["A7"].value or "")
            if not formula.startswith("=FILTER(") or "Base_NCM!" not in formula or "$B$3" not in formula:
                errors.append("Pesquisar!A7 must contain FILTER formula linked to B3 and Base_NCM")
            if str(search["A3"].value or "") != "Pesquisa":
                errors.append("Pesquisar sheet missing visible search label")
        if "Resumo" in wb.sheetnames and wb["Resumo"]["B3"].value != len(rows):
            errors.append("Resumo sheet published row count differs from NDJSON")
    if errors:
        print("\n".join(errors))
        return 1
    print(f"OK: Excel workbook audited with {len(rows)} PIS/Cofins NCM rows: {xlsx}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
